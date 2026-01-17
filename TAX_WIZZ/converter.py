from openpyxl import load_workbook
import json
import os

def num(val):
    try:
        if val is None:
            return 0.0
        if isinstance(val, str):
            return float(val.replace(",", "").strip())
        return float(val)
    except:
        return 0.0


def convert_excel(input_excel, output_dir):
    wb = load_workbook(input_excel, data_only=True, read_only=True)
    sheet = wb.active

    def cell(r, c):
        return sheet.cell(row=r, column=c).value

    example = {
        "Client ID": cell(1, 2),
        "Client Name": cell(2, 2),
        "PAN": cell(3, 2),
        "Taxpnl Statement for Equity from 2024-04-01 to 2025-03-31": [
            {
                "Realized Profit Breakdown": {
                    "Intraday/Speculative profit": num(cell(8, 2)),
                    "Short Term profit": num(cell(9, 2)),
                    "Long Term profit": num(cell(10, 2)),
                    "Non Equity profit": num(cell(11, 2)),
                }
            }
        ]
    }

    example_path = os.path.join(output_dir, "example.json")
    with open(example_path, "w") as f:
        json.dump(example, f, indent=4)

    cg_output = {
        "capitalGain": [],
        "profitLossACIncomes": []
    }

    cg_path = os.path.join(output_dir, "cg_output.json")
    with open(cg_path, "w") as f:
        json.dump(cg_output, f, indent=4)

    wb.close()

    return example, cg_output
