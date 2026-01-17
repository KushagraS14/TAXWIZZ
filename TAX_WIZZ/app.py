"""
Tax Wizz - Excel to JSON Converter
Enhanced Flask Application with Backend-Frontend Sync Features
"""

from flask import Flask, request, jsonify, render_template, redirect, url_for, session, flash, send_from_directory, send_file
from werkzeug.utils import secure_filename
from functools import wraps
from openpyxl import load_workbook
from io import BytesIO, StringIO
import os
import json
import secrets
from datetime import datetime, timedelta
import logging
import threading
import time
import hashlib
from collections import defaultdict

# ============================================
# CONFIGURATION
# ============================================

app = Flask(__name__, static_folder='static', template_folder='templates')

# App Configuration
app.config.update(
    SECRET_KEY=secrets.token_hex(32),
    UPLOAD_FOLDER='uploads',
    CONVERTED_FOLDER='converted_files',
    TEMPLATE_FOLDER='templates',
    MAX_CONTENT_LENGTH=16 * 1024 * 1024,
    ALLOWED_EXTENSIONS={'xlsx', 'xls', 'xlsm', 'xlsb', 'json'},
    SESSION_PERMANENT=False,
    SESSION_COOKIE_SECURE=False,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    # New configurations for sync features
    MAX_HISTORY_PER_USER=50,
    AUTO_SAVE_INTERVAL=300,  # 5 minutes
    BACKUP_RETENTION_DAYS=7
)

# Create necessary directories
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['CONVERTED_FOLDER'], exist_ok=True)
os.makedirs('logs', exist_ok=True)
os.makedirs('user_data', exist_ok=True)

# ============================================
# LOGGING SETUP
# ============================================

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/app.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ============================================
# IN-MEMORY DATA STORES FOR SYNC FEATURES
# ============================================

# User sessions with metadata
USER_SESSIONS = defaultdict(dict)

# File processing queue
PROCESSING_QUEUE = []

# Real-time status updates
STATUS_UPDATES = defaultdict(list)

# User preferences store
USER_PREFERENCES = {}

# Conversion templates
CONVERSION_TEMPLATES = {
    'default': {
        'intraday_start': 42,
        'intraday_end': 42,
        'longterm_start': 55,
        'longterm_end': 57,
        'output_format': 'standard'
    },
    'compact': {
        'intraday_start': 42,
        'intraday_end': 42,
        'longterm_start': 55,
        'longterm_end': 57,
        'output_format': 'compact'
    }
}

# ============================================
# USER DATABASE (Enhanced)
# ============================================

USERS = {
    'admin': {
        'password': 'admin123',
        'name': 'Administrator',
        'email': 'admin@taxwizz.com',
        'role': 'admin',
        'created_at': '2024-01-01',
        'preferences': {
            'theme': 'light',
            'auto_save': True,
            'notifications': True,
            'default_template': 'default'
        }
    },
    'user': {
        'password': 'user123',
        'name': 'Demo User',
        'email': 'user@taxwizz.com',
        'role': 'user',
        'created_at': '2024-01-01',
        'preferences': {
            'theme': 'light',
            'auto_save': True,
            'notifications': True,
            'default_template': 'default'
        }
    }
}

# ============================================
# HELPER FUNCTIONS (Enhanced)
# ============================================

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('login'))
        
        if session.get('role') != 'admin':
            flash('Administrator privileges required.', 'danger')
            return redirect(url_for('index'))
        
        return f(*args, **kwargs)
    return decorated_function

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def num(value):
    try:
        if value is None:
            return 0.0
        if isinstance(value, str):
            cleaned = value.replace(",", "").strip()
            return float(cleaned) if cleaned else 0.0
        return float(value)
    except (ValueError, TypeError):
        return 0.0

def read_excel_data(sheet, start_row, end_row):
    data = []
    for row in range(start_row, end_row + 1):
        symbol = sheet.cell(row=row, column=1).value
        if symbol and str(symbol).strip():
            data.append({
                "Symbol": str(symbol).strip(),
                "Quantity": num(sheet.cell(row=row, column=2).value),
                "Buy Value": num(sheet.cell(row=row, column=3).value),
                "Sell Value": num(sheet.cell(row=row, column=4).value),
                "Realized P&L": num(sheet.cell(row=row, column=5).value),
            })
    return data

def generate_output_filename(original_filename, user_id):
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    name, ext = os.path.splitext(original_filename)
    return f"{user_id}_{name}_{timestamp}.json"

def validate_excel_structure(sheet):
    if sheet.max_row < 2:
        return False, "Excel file appears to be empty"
    return True, "Valid Excel structure"

def get_user_directory(user_id):
    """Get or create user-specific directory"""
    user_dir = os.path.join('user_data', user_id)
    os.makedirs(user_dir, exist_ok=True)
    return user_dir

def log_activity(user_id, activity_type, details):
    """Log user activity for sync and analytics"""
    activity = {
        'timestamp': datetime.now().isoformat(),
        'user_id': user_id,
        'activity_type': activity_type,
        'details': details,
        'ip_address': request.remote_addr
    }
    
    # Store in user session
    if 'activities' not in USER_SESSIONS[user_id]:
        USER_SESSIONS[user_id]['activities'] = []
    
    USER_SESSIONS[user_id]['activities'].append(activity)
    
    # Keep only recent activities
    if len(USER_SESSIONS[user_id]['activities']) > 100:
        USER_SESSIONS[user_id]['activities'] = USER_SESSIONS[user_id]['activities'][-100:]
    
    logger.info(f"Activity: {user_id} - {activity_type}")

def update_user_status(user_id, status, message=""):
    """Update real-time user status for frontend sync"""
    status_update = {
        'timestamp': datetime.now().isoformat(),
        'status': status,
        'message': message
    }
    
    STATUS_UPDATES[user_id].append(status_update)
    
    # Keep only recent status updates
    if len(STATUS_UPDATES[user_id]) > 20:
        STATUS_UPDATES[user_id] = STATUS_UPDATES[user_id][-20:]

def save_user_preferences(user_id, preferences):
    """Save user preferences to disk"""
    user_pref_file = os.path.join(get_user_directory(user_id), 'preferences.json')
    USER_PREFERENCES[user_id] = preferences
    with open(user_pref_file, 'w') as f:
        json.dump(preferences, f, indent=2)

def load_user_preferences(user_id):
    """Load user preferences from disk"""
    user_pref_file = os.path.join(get_user_directory(user_id), 'preferences.json')
    if user_id in USER_PREFERENCES:
        return USER_PREFERENCES[user_id]
    elif os.path.exists(user_pref_file):
        with open(user_pref_file, 'r') as f:
            preferences = json.load(f)
            USER_PREFERENCES[user_id] = preferences
            return preferences
    return USERS.get(user_id, {}).get('preferences', {})

# ============================================
# NEW ROUTES FOR FRONTEND-BACKEND SYNC
# ============================================
@app.route('/')
def home():
    """Home page - redirects to login or index based on auth status"""
    if 'user_id' in session:
        logger.info(f"User {session.get('user_id')} accessing home, redirecting to index")
        return redirect(url_for('index'))
    logger.info("Anonymous user accessing home, redirecting to login")
    return redirect(url_for('login'))

@app.route('/api/sync/status')
@login_required
def sync_status():
    """Get real-time sync status for the user"""
    user_id = session.get('user_id')
    last_sync = USER_SESSIONS[user_id].get('last_sync', datetime.now().isoformat())
    
    return jsonify({
        'online': True,
        'last_sync': last_sync,
        'pending_operations': len(PROCESSING_QUEUE),
        'user_status': USER_SESSIONS[user_id].get('status', 'idle')
    })

@app.route('/api/sync/history')
@login_required
def sync_history():
    """Get user activity history for sync"""
    user_id = session.get('user_id')
    activities = USER_SESSIONS.get(user_id, {}).get('activities', [])
    
    return jsonify({
        'activities': activities[-20:],  # Last 20 activities
        'total_count': len(activities)
    })

@app.route('/api/files/recent')
@login_required
def recent_files():
    """Get recent files for the user"""
    user_id = session.get('user_id')
    user_dir = get_user_directory(user_id)
    
    recent_files = []
    if os.path.exists(user_dir):
        for filename in os.listdir(user_dir):
            if filename.endswith('.json'):
                filepath = os.path.join(user_dir, filename)
                stat = os.stat(filepath)
                recent_files.append({
                    'filename': filename,
                    'size': stat.st_size,
                    'modified': datetime.fromtimestamp(stat.st_mtime).isoformat(),
                    'type': 'json'
                })
    
    return jsonify({'files': sorted(recent_files, key=lambda x: x['modified'], reverse=True)[:10]})

@app.route('/api/preferences', methods=['GET', 'POST', 'PUT'])
@login_required
def user_preferences():
    """Get or update user preferences"""
    user_id = session.get('user_id')
    
    if request.method == 'GET':
        preferences = load_user_preferences(user_id)
        return jsonify(preferences)
    
    elif request.method in ['POST', 'PUT']:
        new_preferences = request.get_json()
        current_preferences = load_user_preferences(user_id)
        
        # Merge preferences
        updated_preferences = {**current_preferences, **new_preferences}
        save_user_preferences(user_id, updated_preferences)
        
        log_activity(user_id, 'preferences_updated', {'preferences': updated_preferences})
        
        return jsonify({
            'success': True,
            'message': 'Preferences updated',
            'preferences': updated_preferences
        })

@app.route('/api/templates', methods=['GET'])
@login_required
def get_templates():
    """Get available conversion templates"""
    return jsonify({
        'templates': CONVERSION_TEMPLATES,
        'default_template': load_user_preferences(session.get('user_id')).get('default_template', 'default')
    })

@app.route('/api/convert/custom', methods=['POST'])
@login_required
def custom_convert():
    """Convert with custom template/parameters"""
    try:
        user_id = session.get('user_id')
        template_name = request.form.get('template', 'default')
        custom_params = request.get_json() if request.is_json else {}
        
        update_user_status(user_id, 'processing', 'Starting custom conversion')
        
        # Get template
        template = CONVERSION_TEMPLATES.get(template_name, CONVERSION_TEMPLATES['default'])
        
        # Apply custom parameters
        if custom_params:
            template.update(custom_params)
        
        # Process file with template
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'error': 'Invalid file type'}), 400
        
        # Process with template parameters
        workbook = load_workbook(filename=BytesIO(file.read()), data_only=True, read_only=True)
        sheet = workbook.active
        
        intraday = read_excel_data(sheet, template['intraday_start'], template['intraday_end'])
        long_term = read_excel_data(sheet, template['longterm_start'], template['longterm_end'])
        
        # Generate output based on template format
        if template['output_format'] == 'compact':
            output_data = generate_compact_output(intraday, long_term)
        else:
            output_data = generate_standard_output(intraday, long_term)
        
        workbook.close()
        
        # Save to user directory
        output_filename = generate_output_filename(file.filename, user_id)
        user_dir = get_user_directory(user_id)
        output_path = os.path.join(user_dir, output_filename)
        
        with open(output_path, 'w') as f:
            json.dump(output_data, f, indent=2)
        
        log_activity(user_id, 'conversion_completed', {
            'template': template_name,
            'filename': file.filename,
            'output_file': output_filename
        })
        
        update_user_status(user_id, 'completed', 'Conversion completed successfully')
        
        return jsonify({
            'success': True,
            'data': output_data,
            'output_file': output_filename,
            'download_url': f'/api/download/{output_filename}'
        })
        
    except Exception as e:
        logger.error(f"Custom conversion error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/download/<filename>')
@login_required
def download_file(filename):
    """Download converted file"""
    user_id = session.get('user_id')
    user_dir = get_user_directory(user_id)
    filepath = os.path.join(user_dir, filename)
    
    if os.path.exists(filepath) and filename.startswith(user_id):
        log_activity(user_id, 'file_downloaded', {'filename': filename})
        return send_file(filepath, as_attachment=True)
    
    return jsonify({'error': 'File not found'}), 404

@app.route('/api/backup', methods=['POST'])
@login_required
def create_backup():
    """Create backup of user data"""
    user_id = session.get('user_id')
    user_dir = get_user_directory(user_id)
    
    # Create backup zip
    import zipfile
    from io import BytesIO
    
    memory_file = BytesIO()
    with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(user_dir):
            for file in files:
                file_path = os.path.join(root, file)
                arcname = os.path.relpath(file_path, user_dir)
                zf.write(file_path, arcname)
    
    memory_file.seek(0)
    
    log_activity(user_id, 'backup_created', {'backup_size': memory_file.getbuffer().nbytes})
    
    return send_file(
        memory_file,
        download_name=f'{user_id}_backup_{datetime.now().strftime("%Y%m%d")}.zip',
        as_attachment=True
    )

@app.route('/api/notifications')
@login_required
def get_notifications():
    """Get user notifications"""
    user_id = session.get('user_id')
    
    # Generate notifications based on activities
    notifications = []
    activities = USER_SESSIONS.get(user_id, {}).get('activities', [])
    
    for activity in activities[-10:]:  # Last 10 activities
        if activity['activity_type'] in ['conversion_completed', 'file_downloaded', 'error_occurred']:
            notifications.append({
                'id': hashlib.md5(json.dumps(activity).encode()).hexdigest()[:8],
                'type': activity['activity_type'],
                'message': f"{activity['activity_type'].replace('_', ' ').title()}: {activity.get('details', {}).get('filename', '')}",
                'timestamp': activity['timestamp'],
                'read': False
            })
    
    return jsonify({'notifications': notifications})

@app.route('/api/stats')
@login_required
def user_stats():
    """Get user statistics"""
    user_id = session.get('user_id')
    user_dir = get_user_directory(user_id)
    
    total_files = 0
    total_size = 0
    if os.path.exists(user_dir):
        for root, dirs, files in os.walk(user_dir):
            total_files += len(files)
            for file in files:
                filepath = os.path.join(root, file)
                total_size += os.path.getsize(filepath)
    
    activities = USER_SESSIONS.get(user_id, {}).get('activities', [])
    conversions = [a for a in activities if a['activity_type'] == 'conversion_completed']
    
    return jsonify({
        'total_files': total_files,
        'total_size': total_size,
        'total_conversions': len(conversions),
        'last_conversion': conversions[-1]['timestamp'] if conversions else None,
        'active_since': USER_SESSIONS[user_id].get('login_time', datetime.now().isoformat())
    })

@app.route('/api/validate/json', methods=['POST'])
@login_required
def validate_json():
    """Validate JSON schema"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'valid': False, 'error': 'No JSON data provided'})
        
        # Basic JSON validation
        import jsonschema
        from jsonschema import validate
        
        # Define a simple schema for validation
        schema = {
            "type": "object",
            "properties": {
                "capitalGain": {"type": "array"},
                "profitLossACIncomes": {"type": "array"}
            },
            "required": ["capitalGain", "profitLossACIncomes"]
        }
        
        validate(instance=data, schema=schema)
        return jsonify({'valid': True, 'message': 'JSON is valid'})
        
    except jsonschema.ValidationError as e:
        return jsonify({'valid': False, 'error': str(e)})
    except Exception as e:
        return jsonify({'valid': False, 'error': str(e)})

# ============================================
# ENHANCED EXISTING ROUTES
# ============================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Enhanced login with sync features"""
    if 'user_id' in session:
        return redirect(url_for('index'))
    
    error = None
    
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        remember = request.form.get('remember')
        
        if not username or not password:
            error = 'Please enter both username and password'
        elif username in USERS and USERS[username]['password'] == password:
            # Login successful
            session['user_id'] = username
            session['user_name'] = USERS[username]['name']
            session['user_email'] = USERS[username]['email']
            session['role'] = USERS[username]['role']
            session['logged_in'] = True
            
            # Initialize user session
            USER_SESSIONS[username] = {
                'login_time': datetime.now().isoformat(),
                'ip_address': request.remote_addr,
                'status': 'active',
                'last_sync': datetime.now().isoformat()
            }
            
            log_activity(username, 'login', {'ip': request.remote_addr})
            
            if remember:
                session.permanent = True
            
            flash('Login successful! Welcome to Tax Wizz.', 'success')
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('index'))
        else:
            error = 'Invalid username or password'
    
    return render_template('login.html', error=error)

@app.route('/convert', methods=['POST'])
@login_required
def convert():
    """Enhanced convert with sync features"""
    try:
        user_id = session.get('user_id')
        
        # Update user status
        update_user_status(user_id, 'uploading', 'Receiving file')
        
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'error': 'Invalid file type'}), 400
        
        filename = secure_filename(file.filename)
        
        update_user_status(user_id, 'processing', 'Reading Excel file')
        log_activity(user_id, 'file_uploaded', {'filename': filename})
        
        # Process file
        workbook = load_workbook(filename=BytesIO(file.read()), data_only=True, read_only=True)
        sheet = workbook.active
        
        is_valid, validation_message = validate_excel_structure(sheet)
        if not is_valid:
            workbook.close()
            return jsonify({'success': False, 'error': validation_message}), 400
        
        update_user_status(user_id, 'processing', 'Extracting data')
        
        # Read data
        intraday = read_excel_data(sheet, 42, 42)
        long_term = read_excel_data(sheet, 55, 57)
        
        # Generate output
        update_user_status(user_id, 'processing', 'Generating JSON output')
        response_data = generate_standard_output(intraday, long_term)
        
        workbook.close()
        
        # Save to user directory
        output_filename = generate_output_filename(filename, user_id)
        user_dir = get_user_directory(user_id)
        output_path = os.path.join(user_dir, output_filename)
        
        with open(output_path, 'w') as f:
            json.dump(response_data, f, indent=2)
        
        log_activity(user_id, 'conversion_completed', {
            'filename': filename,
            'output_file': output_filename,
            'intraday_count': len(intraday),
            'longterm_count': len(long_term)
        })
        
        update_user_status(user_id, 'completed', 'Conversion successful')
        
        return jsonify({
            'success': True,
            'data': response_data,
            'message': 'File processed successfully!',
            'output_file': output_filename,
            'download_url': f'/api/download/{output_filename}',
            'stats': {
                'intraday_trades': len(intraday),
                'long_term_trades': len(long_term)
            }
        })
        
    except Exception as e:
        logger.error(f"Conversion error: {str(e)}")
        user_id = session.get('user_id')
        log_activity(user_id, 'error_occurred', {'error': str(e)})
        update_user_status(user_id, 'error', str(e))
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/index')
@login_required
def index():
    """Enhanced index with sync data"""
    user_id = session.get('user_id')
    
    # Get user preferences
    preferences = load_user_preferences(user_id)
    
    # Get recent activities
    activities = USER_SESSIONS.get(user_id, {}).get('activities', [])[-5:]
    
    # Get user stats
    user_dir = get_user_directory(user_id)
    file_count = len([f for f in os.listdir(user_dir) if f.endswith('.json')]) if os.path.exists(user_dir) else 0
    
    user_info = {
        'username': user_id,
        'name': session.get('user_name'),
        'email': session.get('user_email'),
        'role': session.get('role'),
        'preferences': preferences,
        'stats': {
            'file_count': file_count,
            'last_activity': activities[0]['timestamp'] if activities else None
        }
    }
    
    return render_template('index.html', 
                         user=user_info,
                         recent_activities=activities)

# ============================================
# HELPER FUNCTIONS FOR OUTPUT GENERATION
# ============================================

def generate_standard_output(intraday, long_term):
    """Generate standard output format"""
    capital_gain = []
    if long_term:
        asset_details = []
        for srn, trade in enumerate(long_term, start=1):
            qty = trade["Quantity"]
            sell_value_per_unit = trade["Sell Value"] / qty if qty else 0
            purchase_value_per_unit = trade["Buy Value"] / qty if qty else 0
            
            asset_details.append({
                "srn": srn,
                "gainType": "LONG",
                "sellDate": "2025-03-31T18:30:00Z",
                "purchaseDate": "2024-04-01T18:30:00Z",
                "sellValue": trade["Sell Value"],
                "purchaseCost": trade["Buy Value"],
                "sellValuePerUnit": sell_value_per_unit,
                "purchaseValuePerUnit": purchase_value_per_unit,
                "sellOrBuyQuantity": qty,
                "nameOfTheUnits": trade["Symbol"],
                "capitalGain": trade["Realized P&L"],
                "algorithm": "cgSharesMF",
                "brokerName": "Manual"
            })
        
        capital_gain.append({
            "assessmentYear": "2025-2026",
            "assesseeType": "INDIVIDUAL",
            "assetType": "EQUITY_SHARES_LISTED",
            "assetDetails": asset_details
        })
    
    profit_loss = []
    if intraday:
        total_pnl = sum(t["Realized P&L"] for t in intraday)
        total_turnover = sum(t["Sell Value"] for t in intraday)
        
        profit_loss.append({
            "businessType": "SPECULATIVEINCOME",
            "netProfitfromSpeculativeIncome": total_pnl,
            "incomes": [
                {
                    "turnOver": total_turnover,
                    "grossProfit": total_pnl,
                    "brokerName": "Manual"
                }
            ]
        })
    
    return {
        "capitalGain": capital_gain,
        "profitLossACIncomes": profit_loss,
        "metadata": {
            "generated_at": datetime.now().isoformat(),
            "version": "2.0",
            "format": "standard"
        }
    }

def generate_compact_output(intraday, long_term):
    """Generate compact output format"""
    output = {
        "summary": {
            "intraday_trades": len(intraday),
            "long_term_trades": len(long_term),
            "total_intraday_pnl": sum(t["Realized P&L"] for t in intraday),
            "total_longterm_pnl": sum(t["Realized P&L"] for t in long_term),
            "generated_at": datetime.now().isoformat()
        },
        "trades": {
            "intraday": intraday,
            "long_term": long_term
        }
    }
    return output

# ============================================
# BACKGROUND TASKS (Threading)
# ============================================

@app.errorhandler(404)
def page_not_found(e):
    """Handle 404 errors"""
    logger.warning(f"404 error: {request.url}")
    return render_template('error.html', 
                         error_code=404,
                         error_message='Page not found'), 404

@app.errorhandler(500)
def internal_server_error(e):
    """Handle 500 errors"""
    logger.error(f"500 error: {str(e)}")
    return render_template('error.html',
                         error_code=500,
                         error_message='Internal server error'), 500

@app.errorhandler(413)
def too_large(e):
    """Handle file too large errors"""
    logger.warning(f"File too large: {request.remote_addr}")
    return jsonify({
        'success': False,
        'error': f'File too large. Maximum size is {app.config["MAX_CONTENT_LENGTH"] / (1024*1024)}MB'
    }), 413

# ============================================
# UTILITY ROUTES
# ============================================
@app.route('/logout')
@login_required
def logout():
    user_id = session.get('user_id')

    if user_id:
        log_activity(user_id, 'logout', {'ip': request.remote_addr})
        USER_SESSIONS.pop(user_id, None)

    session.clear()
    flash('You have been logged out successfully.', 'info')
    return redirect(url_for('login'))



@app.route('/health')
def health_check():
    """Health check endpoint for monitoring"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat(),
        'version': '1.0.0'
    })

@app.route('/favicon.ico')
def favicon():
    """Serve favicon"""
    return send_from_directory(os.path.join(app.root_path, 'static'),
                             'favicon.ico', mimetype='image/vnd.microsoft.icon')

# ============================================
# APPLICATION STARTUP
# ============================================

if __name__ == '__main__':
    print("=" * 60)
    print("🚀 TAX WIZZ - Enhanced Excel to JSON Converter")
    print("=" * 60)
    print("📁 User data folder: user_data/")
    print("📁 Upload folder: uploads/")
    print("📁 Converted files: converted_files/")
    print("=" * 60)
    print("🔄 Sync Features Enabled:")
    print("   • Real-time status updates")
    print("   • User activity tracking")
    print("   • File history management")
    print("   • User preferences sync")
    print("   • Automatic cleanup")
    print("=" * 60)
    print(f"🌐 Server starting on: http://localhost:5000")
    print("=" * 60)
    
    app.run(
        host='0.0.0.0',
        port=5000,
        debug=True,
        use_reloader=True
    )